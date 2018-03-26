import TSP as FileRead, openpyxl as op, os
from datetime import datetime as dt
from openpyxl import workbook

if os.path.isfile("ProjectExcel.xlsx") == False:
    wb = op.Workbook("ProjectExcel.xlsx",data_only=True)
    ws = wb.create_sheet(title="Report")
    ws.append(['Time (s)','Problem Type','Name','Dimension','Solution','Optimality Gap','Optimal Solution','Vehicles','Number of Iterations','BMU Neighborhood'])
    wb.save("ProjectExcel.xlsx")

wb = op.load_workbook("ProjectExcel.xlsx",data_only=True)
ws = wb["Report"]

ws.cell(row=ws.max_row+1,column=1,value=FileRead.CodeElapsedTime)

ws.cell(row=ws.max_row,column=2,value=FileRead.problemtype)

ws.cell(row=ws.max_row,column=3,value=FileRead.filename)

ws.cell(row=ws.max_row,column=4,value=FileRead.dimension)

ws.cell(row=ws.max_row,column=5,value=FileRead.totalDistance)

ws.cell(row=ws.max_row,column=6,value=FileRead.OptGap)

ws.cell(row=ws.max_row,column=7,value=FileRead.optcost)

ws.cell(row=ws.max_row,column=8,value=int(1))

ws.cell(row=ws.max_row,column=9,value=FileRead.NumIters)

ws.cell(row=ws.max_row,column=10,value=FileRead.BMU_effect)

wb.save("ProjectExcel.xlsx")
